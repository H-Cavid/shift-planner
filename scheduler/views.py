from django.shortcuts import render, redirect,HttpResponse
from django.contrib.auth import login
from .forms import CustomUserCreationForm  # use the custom form



def register_view(request):
    if request.method == "POST":
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.role = 'worker'  # auto-assign 'worker' role
            user.save()
            login(request, user)
            return redirect('/accounts/login/')  # temporary redirect
    else:
        form = CustomUserCreationForm()
    return render(request, "registration/register.html", {"form": form})


from django.contrib.auth.decorators import login_required
from django.shortcuts import render

# @login_required
# def home_view(request):
#     return render(request, "home.html", {"user": request.user})

# @login_required
# def home_view(request):
#     if request.user.role == 'manager':
#         return HttpResponse("👔 Welcome Manager")  # Will later change to manager dashboard
#     else:
#         return redirect('my_shifts')  # Redirect workers to their shifts


from django.shortcuts import render
from django.contrib.auth.decorators import login_required

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect

@login_required
def home_view(request):
    if request.user.role == 'manager' or request.user.is_superuser:
        return redirect('manager_dashboard')
    return render(request, 'home.html')

# def home_view(request):
#     # Send managers to their dashboard; workers to their home (or shifts)
#     if request.user.role == 'manager':
#         return redirect('manager_dashboard')
#     return render(request, 'home.html')   # worker landing page


# @login_required
# def manager_dashboard(request):
#     return render(request, 'manager/dashboard.html')

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required


# bu asagidaki isiyirdi
# @login_required
# def manager_dashboard(request):
#     if request.user.role != 'manager':
#         return redirect('home')
#     return render(request, 'manager_dashboard.html')  # make sure this is correct

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Shift

# bu hisse isdiyir graphic gosterir
# @login_required
# def manager_dashboard(request):
#     if request.user.role != 'manager':
#         return redirect('home')

#     User = get_user_model()
#     workers = User.objects.filter(role='worker')

#     worker_names = []
#     total_paid_hours = []

#     for worker in workers:
#         shifts = Shift.objects.filter(worker=worker)
#         total = sum([s.paid_hours() for s in shifts])
#         worker_names.append(worker.username)
#         total_paid_hours.append(round(total, 1))

#     return render(request, 'manager_dashboard.html', {
#         'worker_names': worker_names,
#         'paid_hours': total_paid_hours,
#     })

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import Shift
from collections import defaultdict

# bu asagidaki isdeyir qrafiknen textnen bir yerde

# @login_required
# def manager_dashboard(request):
#     if request.user.role != 'manager':
#         return redirect('home')

#     workers_data = defaultdict(lambda: [])

#     for shift in Shift.objects.all().order_by('worker__username', 'date'):
#         duration = shift.paid_hours()
#         workers_data[shift.worker.username].append({
#             'date': shift.date.strftime('%Y-%m-%d'),
#             'paid': duration
#         })

#     return render(request, 'manager_dashboard.html', {
#         'workers_data': dict(workers_data)
#     })

from collections import defaultdict
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Shift
from django.utils.dateparse import parse_date
# bu asagidaki isleyir gunlernen bir yerde
# @login_required 
# def manager_dashboard(request):
#     if request.user.role != 'manager':
#         return redirect('home')

#     # 📊 Existing chart logic
#     workers_data = defaultdict(lambda: [])
#     for shift in Shift.objects.all().order_by('worker__username', 'date'):
#         duration = shift.paid_hours()
#         workers_data[shift.worker.username].append({
#             'date': shift.date.strftime('%Y-%m-%d'),
#             'paid': duration
#         })

#     # 📅 New: Date filter logic
#     date_filter = request.GET.get('date')
#     shifts_on_date = []
#     if date_filter:
#         parsed_date = parse_date(date_filter)
#         shifts_on_date = Shift.objects.filter(date=parsed_date).select_related('worker')

#     return render(request, 'manager_dashboard.html', {
#         'workers_data': dict(workers_data),
#         'shifts_on_date': shifts_on_date,
#         'date_filter': date_filter  # Pass it back to fill the input
#     })

from django.contrib.auth import get_user_model
from collections import defaultdict
from datetime import datetime

@login_required 
def manager_dashboard(request):
    if request.user.role != 'manager':
        return redirect('home')

    User = get_user_model()
    date_filter = request.GET.get('date')  # e.g. '2025-08-01'
    shifts_on_date = []
    free_workers = []

    if date_filter:
        shifts_on_date = Shift.objects.filter(date=date_filter).select_related('worker')
        working_users = [shift.worker for shift in shifts_on_date]
        free_workers = User.objects.filter(role='worker').exclude(id__in=[user.id for user in working_users])
    else:
        shifts_on_date = []

    # Bar chart data
    workers_data = defaultdict(list)
    for shift in Shift.objects.all().order_by('worker__username', 'date'):
        workers_data[shift.worker.username].append({
            'date': shift.date.strftime('%Y-%m-%d'),
            'paid': shift.paid_hours()
        })

    return render(request, 'manager_dashboard.html', {
        'workers_data': dict(workers_data),
        'shifts_on_date': shifts_on_date,
        'free_workers': free_workers,
        'date_filter': date_filter,
    })




# # scheduler/views.py
# from django.shortcuts import render
# from .models import Shift

# @login_required
# def filter_by_date(request):
#     workers_on_date = []
#     selected_date = None

#     if request.method == 'POST':
#         selected_date = request.POST.get('shift_date')
#         if selected_date:
#             shifts = Shift.objects.filter(date=selected_date)
#             workers_on_date = [shift.worker for shift in shifts]

#     return render(request, 'filter_by_date.html', {
#         'workers': workers_on_date,
#         'selected_date': selected_date
#     })




from django.contrib.auth.decorators import login_required
from .models import Shift

# @login_required
# def my_shifts_view(request):
#     if request.user.role != 'worker':
#         return redirect('home')  # or raise permission error

#     shifts = Shift.objects.filter(worker=request.user).order_by('date', 'start_time')
#     return render(request, 'my_shifts.html', {'shifts': shifts})

@login_required
def my_shifts_view(request):
    if request.user.role != 'worker':
        return redirect('home')

    shifts = Shift.objects.filter(worker=request.user).order_by('date', 'start_time')

    total_paid = sum(shift.paid_hours() for shift in shifts)
    total_duration = sum(shift.duration_hours() for shift in shifts)

    return render(request, 'my_shifts.html', {
        'shifts': shifts,
        'total_paid': total_paid,
        'total_duration': total_duration
    })




###
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from .models import Availability
from .forms import AvailabilityForm

# 🔹 List all availability entries for the current user
class MyAvailabilityListView(LoginRequiredMixin, ListView):
    model = Availability
    template_name = 'availability/my_availabilities.html'
    context_object_name = 'availabilities'

    def get_queryset(self):
        return Availability.objects.filter(worker=self.request.user)






# 🔹 Create availability
class AvailabilityCreateView(LoginRequiredMixin, CreateView):
    model = Availability
    form_class = AvailabilityForm
    template_name = 'availability/availability_form.html'
    success_url = reverse_lazy('my_availability')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['worker'] = self.request.user  # ✅ pass to form
        return kwargs

    def form_valid(self, form):
        form.instance.worker = self.request.user  # ✅ assign before saving
        return super().form_valid(form)


# class AvailabilityCreateView(LoginRequiredMixin, CreateView):
#     model = Availability
#     form_class = AvailabilityForm
#     template_name = 'availability/availability_form.html'
#     success_url = reverse_lazy('my_availability')

#     def form_valid(self, form):
#         form.instance.worker = self.request.user
#         return super().form_valid(form)

# class AvailabilityCreateView(LoginRequiredMixin, CreateView):
#     model = Availability
#     form_class = AvailabilityForm
#     template_name = 'availability/availability_form.html'  # <-- match your template path
#     success_url = reverse_lazy('my_availability')

#     # def get_form_kwargs(self):
#     #     kwargs = super().get_form_kwargs()
#     #     kwargs['worker'] = self.request.user   # pass user to the form for validation
#     #     return kwargs

#     def form_valid(self, form):
#         form.instance.worker = self.request.user
#         return super().form_valid(form)


# 🔹 Edit only own availability
# class AvailabilityUpdateView(LoginRequiredMixin, UpdateView):
#     model = Availability
#     form_class = AvailabilityForm
#     template_name = 'availability/availability_form.html'
#     success_url = reverse_lazy('my_availability')

#     def get_queryset(self):
#         return Availability.objects.filter(worker=self.request.user)

class AvailabilityUpdateView(LoginRequiredMixin, UpdateView):
    model = Availability
    form_class = AvailabilityForm
    template_name = 'availability/availability_form.html'
    success_url = reverse_lazy('my_availability')

    def get_queryset(self):
        return Availability.objects.filter(worker=self.request.user)

    # def get_form_kwargs(self):
    #     kwargs = super().get_form_kwargs()
    #     kwargs['worker'] = self.request.user
    #     return kwargs


# 🔹 Delete only own availability
class AvailabilityDeleteView(LoginRequiredMixin, DeleteView):
    model = Availability
    template_name = 'availability/availability_confirm_delete.html'
    success_url = reverse_lazy('my_availability')

    def get_queryset(self):
        return Availability.objects.filter(worker=self.request.user)


from django.forms import modelformset_factory
from django.contrib.admin.views.decorators import staff_member_required
from .models import Shift
from .forms import ShiftForm

@staff_member_required  # only manager/superuser can access
def assign_multiple_shifts_view(request):
    ShiftFormSet = modelformset_factory(Shift, form=ShiftForm, extra=5, can_delete=False)

    if request.method == 'POST':
        formset = ShiftFormSet(request.POST)
        if formset.is_valid():
            instances = formset.save(commit=False)
            for shift in instances:
                shift.created_by = request.user
                shift.save()
            return redirect('home')
    else:
        formset = ShiftFormSet(queryset=Shift.objects.none())  # show empty forms

    return render(request, 'assign_multiple_shifts.html', {'formset': formset})



# from datetime import datetime  # ✅ this gives you direct access to datetime.combine
# import openpyxl
# from django.http import HttpResponse
# from .models import Shift

# def export_shifts_excel(request):
#     if not request.user.is_authenticated or request.user.role != 'manager':
#         return HttpResponse(status=403)

#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Shifts Report"

#     ws.append(["Worker", "Date", "Start Time", "End Time", "Duration (hours)", "Paid Hours (€)"])

#     for shift in Shift.objects.all().order_by('worker', 'date'):
#         start_dt = datetime.combine(datetime.today(), shift.start_time)
#         end_dt = datetime.combine(datetime.today(), shift.end_time)
#         duration = (end_dt - start_dt).total_seconds() / 3600

#         ws.append([
#             shift.worker.username,
#             shift.date.strftime('%Y-%m-%d'),
#             shift.start_time.strftime('%H:%M'),
#             shift.end_time.strftime('%H:%M'),
#             round(duration, 2),
#             round(shift.paid_hours(), 2)  # ✅ Add paid hours
#         ])

#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = 'attachment; filename=shifts_report.xlsx'
#     wb.save(response)
#     return response

from collections import defaultdict
from datetime import datetime
import openpyxl
from django.http import HttpResponse
from .models import Shift

def export_shifts_excel(request):
    if not request.user.is_authenticated or request.user.role != 'manager':
        return HttpResponse(status=403)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shifts Report"

    # Header
    ws.append(["Worker", "Date", "Start Time", "End Time", "Duration (hours)", "Paid Hours (€)"])

    # Group shifts by worker
    worker_shifts = defaultdict(list)
    for shift in Shift.objects.all().order_by('worker__username', 'date'):
        worker_shifts[shift.worker.username].append(shift)

    for worker, shifts in worker_shifts.items():
        total_hours = 0
        total_paid = 0

        for shift in shifts:
            start_dt = datetime.combine(datetime.today(), shift.start_time)
            end_dt = datetime.combine(datetime.today(), shift.end_time)
            duration = (end_dt - start_dt).total_seconds() / 3600
            paid = shift.paid_hours()

            total_hours += duration
            total_paid += paid

            ws.append([
                worker,
                shift.date.strftime('%Y-%m-%d'),
                shift.start_time.strftime('%H:%M'),
                shift.end_time.strftime('%H:%M'),
                round(duration, 2),
                round(paid, 2)
            ])

        # Add total row for this worker
        ws.append([
            f"Total for {worker}", "", "", "",
            round(total_hours, 2),
            round(total_paid, 2)
        ])

        # Empty row between workers
        ws.append([])

    # Generate response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=shifts_report.xlsx'
    wb.save(response)
    return response


# from django.template.loader import get_template
# from xhtml2pdf import pisa
# from django.http import HttpResponse
# from io import BytesIO

# def export_shifts_pdf(request):
#     if not request.user.is_authenticated or request.user.role != 'manager':
#         return HttpResponse(status=403)

#     shifts = Shift.objects.all().order_by('worker__username', 'date')

#     template_path = 'pdf_shifts_report.html'
#     context = {'shifts': shifts}

#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = 'attachment; filename="shifts_report.pdf"'

#     template = get_template(template_path)
#     html = template.render(context)

#     pisa_status = pisa.CreatePDF(html, dest=response)

#     if pisa_status.err:
#         return HttpResponse('Error creating PDF', status=500)
#     return response

from collections import defaultdict
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
from datetime import datetime

def export_shifts_pdf(request):
    if not request.user.is_authenticated or request.user.role != 'manager':
        return HttpResponse(status=403)

    shifts_by_worker = defaultdict(lambda: {'shifts': [], 'total_hours': 0, 'total_paid': 0})

    for shift in Shift.objects.all().order_by('worker__username', 'date'):
        start_dt = datetime.combine(datetime.today(), shift.start_time)
        end_dt = datetime.combine(datetime.today(), shift.end_time)
        duration = (end_dt - start_dt).total_seconds() / 3600
        paid = shift.paid_hours()

        shifts_by_worker[shift.worker.username]['shifts'].append({
            'date': shift.date,
            'start': shift.start_time.strftime('%H:%M'),
            'end': shift.end_time.strftime('%H:%M'),
            'duration': round(duration, 1),
            'paid': round(paid, 1),
        })

        shifts_by_worker[shift.worker.username]['total_hours'] += duration
        shifts_by_worker[shift.worker.username]['total_paid'] += paid

    context = {
        'shifts_by_worker': shifts_by_worker
    }

    template = get_template('pdf_shifts_report.html')
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="shifts_report.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('PDF generation error', status=500)
    return response



# import openpyxl
# from django.http import HttpResponse
# from .models import Shift
# from datetime import datetime

# def export_my_shifts_excel(request):
#     user = request.user

#     # Only allow workers to export their own shifts
#     if not user.is_authenticated or user.role != 'worker':
#         return HttpResponse("Unauthorized", status=401)

#     shifts = Shift.objects.filter(worker=user).order_by('date')

#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "My Shifts"

#     # Header row
#     ws.append(['Date', 'Start', 'End', 'Raw Duration (h)', 'Paid Hours (h)'])

#     # Add shift rows
#     for shift in shifts:
#         raw_duration = round((datetime.combine(shift.date, shift.end_time) - datetime.combine(shift.date, shift.start_time)).seconds / 3600, 1)
#         ws.append([
#             shift.date.strftime('%Y-%m-%d'),
#             shift.start_time.strftime('%H:%M'),
#             shift.end_time.strftime('%H:%M'),
#             raw_duration,
#             shift.paid_hours()
#         ])


#     # Create response
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#     )
#     response['Content-Disposition'] = f'attachment; filename={user.username}_shifts.xlsx'
#     wb.save(response)
#     return response

import openpyxl
from django.http import HttpResponse
from .models import Shift
from datetime import datetime

def export_my_shifts_excel(request):
    user = request.user

    # Only allow workers to export their own shifts
    if not user.is_authenticated or user.role != 'worker':
        return HttpResponse("Unauthorized", status=401)

    shifts = Shift.objects.filter(worker=user).order_by('date')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Shifts"

    # Header row
    ws.append(['Date', 'Start', 'End', 'Raw Duration (h)', 'Paid Hours (h)'])

    total_paid = 0  # Track total paid hours

    # Add shift rows
    for shift in shifts:
        raw_duration = round((datetime.combine(shift.date, shift.end_time) - datetime.combine(shift.date, shift.start_time)).seconds / 3600, 1)
        paid = shift.paid_hours()
        total_paid += paid
        ws.append([
            shift.date.strftime('%Y-%m-%d'),
            shift.start_time.strftime('%H:%M'),
            shift.end_time.strftime('%H:%M'),
            raw_duration,
            paid
        ])

    # Add an empty row before totals
    ws.append([])

    # Add total paid hours row
    ws.append(['Total Paid Hours (Planstunden):', '', '', '', round(total_paid, 1)])

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={user.username}_shifts.xlsx'
    wb.save(response)
    return response


# from django.template.loader import get_template
# from django.http import HttpResponse
# from xhtml2pdf import pisa
# from .models import Shift

# def export_my_shifts_pdf(request):
#     user = request.user

#     if not user.is_authenticated or user.role != 'worker':
#         return HttpResponse("Unauthorized", status=401)

#     shifts = Shift.objects.filter(worker=user).order_by('date')

#     # Group shifts by worker (even though it's one user here, keeping it clean)
#     shifts_by_worker = {user: shifts}

#     # Calculate total paid hours
#     total_paid_hours = sum([shift.paid_hours() for shift in shifts])

#     # Load template
#     template = get_template("pdf_shifts_report.html")
#     html = template.render({
#         "shifts_by_worker": shifts_by_worker,
#         "total_paid_hours": round(total_paid_hours, 1),
#     })

#     # Create PDF response
#     response = HttpResponse(content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename={user.username}_shifts.pdf'
#     pisa_status = pisa.CreatePDF(html, dest=response)

#     if pisa_status.err:
#         return HttpResponse("Error generating PDF", status=500)
#     return response

from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from .models import Shift

def export_my_shifts_pdf(request):
    user = request.user

    if not user.is_authenticated or user.role != 'worker':
        return HttpResponse("Unauthorized", status=401)

    shifts = Shift.objects.filter(worker=user).order_by('date')

    # Prepare a single worker structure (no need to loop in template anymore)
    shift_data = [
        {
            "date": shift.date,
            "start": shift.start_time.strftime('%H:%M'),
            "end": shift.end_time.strftime('%H:%M'),
            "duration": shift.duration_hours(),
            "paid": shift.paid_hours(),
        }
        for shift in shifts
    ]

    total_hours = sum([s["duration"] for s in shift_data])
    total_paid = sum([s["paid"] for s in shift_data])

    template = get_template("pdf_shifts_report.html")
    html = template.render({
        "user": user,
        "shifts": shift_data,
        "total_hours": round(total_hours, 1),
        "total_paid": round(total_paid, 1),
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={user.username}_shifts.pdf'
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse("PDF generation error", status=500)

    return response


